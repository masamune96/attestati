import os
import openpyxl
import asyncio
from tqdm import tqdm
from utils.style import *
from utils.format import Formatter
from utils.config import pdf_folder, excel_path
from gpt.tokens import shared
from gpt.openai_api import openai_call, crea_batch

# numero massimo di batch che possono essere elaborati al minuto
# Token limits Tier2 gpt-4o: 450.000 TPM (Tokens per minute)
# con Tier2 800.000 TPM (posso aumentare i MAX_BATCHES_PER_MINUTE)
MAX_BATCHES_PER_MINUTE = 4

async def process_batch(batch, batch_size, sheet, row_counter, lock):
    try:
        analisi_batch = await openai_call([(t["nome_file"], t["testo"]) for t in batch], batch_size)

        # allinea i JSON agli attestati nel batch
        json_completi = Formatter.allinea_json(batch, analisi_batch)

        async with lock:  # assicura che scriva solo un thread alla volta
            for i, attestato in enumerate(json_completi):
                try:
                    nome_file = batch[i]["nome_file"]
                    testo_attestato = batch[i]["testo"]
                    nome_corso = Formatter.pulisci_nome_corso(attestato.get("nome_corso", "ND")).upper()

                    # scrittura nell'Excel
                    sheet[f'B{row_counter[0]}'] = attestato.get("nome_partecipante", "ND").upper()
                    sheet[f'C{row_counter[0]}'] = attestato.get("cognome_partecipante", "ND").upper()
                    sheet[f'D{row_counter[0]}'] = attestato.get("data_fine_corso", "ND")
                    sheet[f'E{row_counter[0]}'] = attestato.get("dati_anagrafici", "ND")
                    sheet[f'F{row_counter[0]}'] = Formatter.valida_cf(attestato.get("codice_fiscale", "ND"))
                    sheet[f'I{row_counter[0]}'] = attestato.get("tdi", "ND")
                    sheet[f'J{row_counter[0]}'] = nome_file
                    sheet[f'H{row_counter[0]}'] = attestato.get("codice_corso", "ND")
                    # DEBUG columns
                    sheet[f'T{row_counter[0]}'] = nome_corso
                    sheet[f'U{row_counter[0]}'] = testo_attestato
                    sheet[f'V{row_counter[0]}'] = str(attestato)

                    row_counter[0] += 1  # aggiorna la riga in modo thread-safe

                except Exception as e:
                    print(f"{RED}Errore durante la scrittura dei dati per {nome_file}: {e}{RESET}")

    except Exception as e:
        print(f"{RED}Errore durante l'elaborazione del batch: {e}{RESET}")
        raise RuntimeError(f"Errore critico durante l'elaborazione del batch: {e}")

# gestisce il rate-limiting globale per limitare a 3 batch al minuto
async def rate_limiter():
    while True:
        await asyncio.sleep(60)  # aspetta un minuto per nuovi batch
        for _ in range(MAX_BATCHES_PER_MINUTE):
            batch_semaphore.release()

async def aggiorna_excel(excel_path, pdf_folder):
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # cerca prima riga con nome e cognome vuoti
        row = 1
        for current_row in range(1, sheet.max_row + 1):
            if (not sheet[f'B{current_row}'].value) and (not sheet[f'C{current_row}'].value):
                row = current_row
                break
        else:
            row = sheet.max_row + 1  

        # crea dizionario attestati con nome file, testo, num token per ogni attestato
        testi_estratti = shared.token_per_pdf(pdf_folder)
        # crea list di batch da inviare all'API
        batches = list(crea_batch(testi_estratti))
        # Lock asincrono
        row_counter = [row]
        lock = asyncio.Lock()  

        # semaforo per limitare i batch (limite di tokens al minuto)
        global batch_semaphore
        batch_semaphore = asyncio.Semaphore(MAX_BATCHES_PER_MINUTE)
        # avvia il rate-limiter in background
        asyncio.create_task(rate_limiter())

        print(f"{BOLD}{CYAN}Richiesta API in corso...{RESET} ⏳")

        # elabora i batch in parallelo con tasks
        tasks = [process_batch(batch, batch_size, sheet, row_counter, lock) for batch_size, batch in batches]

        for f in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="Elaborazione Batch"):
            await batch_semaphore.acquire()  # aspetta disponibilità dal rate limiter
            await f

        workbook.save(excel_path)
        print(f"\nDati aggiornati in {BOLD}{YELLOW}{excel_path}{RESET} \u2705\n")

    except Exception as e:
        print(f"{RED}{BOLD}Errore durante l'aggiornamento del file Excel:{RESET} {e}")





if __name__ == "__main__":
    if not os.path.exists(pdf_folder):
        print(f"{RED}{BOLD}Cartella PDF {pdf_folder} non trovata.{RESET}")
    else:
        asyncio.run(aggiorna_excel(excel_path, pdf_folder))
        print(f"\U0001F9E9 Token totali: {BOLD}{shared.token_totali} ({shared.token_totali_input} input + {shared.token_totali_output} output){RESET}")
        print(f"\U0001F4B8 Costo totale: {BOLD}{shared.price()} \u20AC{RESET}")